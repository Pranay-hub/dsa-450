#!/usr/bin/env python3
"""Generate complete DSA 450 problems.json from spreadsheet data"""
import json, re, os

def slug(name):
    s = re.sub(r'[^a-z0-9]+', '-', name.lower().strip()).strip('-')
    return s

# Read spreadsheet data
import openpyxl
wb = openpyxl.load_workbook('/Users/owner/Downloads/1787154816695_DSA450.xlsx')
ws = wb.active

raw = []
for row in range(6, ws.max_row+1):
    c1 = (ws.cell(row=row, column=1).value or '').strip()
    c2 = (ws.cell(row=row, column=2).value or '').strip()
    c3 = (ws.cell(row=row, column=3).value or '').strip()
    c4 = (ws.cell(row=row, column=4).value or '').strip()
    if c2:
        raw.append((row, c1, c2, c3, c4))

def status_map(s):
    if s == 'yes': return 'solved'
    if s == 'refer later' or s == 'refer code later' or s == 'refer later again' or s == 'refer again after dp' or s == 'refer later again': return 'pending'
    if s == '<->': return 'pending'
    return 'pending'

def get_difficulty(topic, name, row):
    hard_problems = {
        'Count Inversion', 'Trapping Rain water problem', 'Edit Distance',
        'Longest Palindromic Substring', 'N-Queen Problem', 'Sudoku Solver',
        'Knight\'s tour problem', 'Tug of War', 'Matrix Chain Multiplication',
        'Egg Dropping Problem', 'Boolean Parenthesization Problem',
        'Travelling Salesman Problem', 'Maze', 'Word Break Problem using Backtracking',
        'Remove Invalid Parentheses', 'Longest Possible Route in a Matrix with Hurdles',
        'Longest path in a Directed Acyclic Graph', 'Cheapest Flights Within K Stops',
        'Find bridge in a graph', 'Largest BST in a Binary Tree',
        'Convert a normal BST into a Balanced BST', 'Merge two BST',
        'Count Strongly connected Components(Kosaraju Algo)', 'Median in a row-wise sorted Matrix',
        'Find Median of Two Sorted Arrays', 'Median in a stream of Integers',
        'Smallest range in K Lists', 'Word Ladder', 'Implement Floyd warshall Algorithm',
        'Implement Bellman Ford Algorithm', 'Find shortest safe route in a path with landmines',
        'Find Maximum number possible by doing at-most K swaps',
        'Partition of a set into K subsets with equal sum', 'Length of the Longest Valid Substring',
        'Largest rectangular Area in Histogram', 'Minimum Window Substring',
        'Longest Repeating Subsequence', 'Minimize the maximum difference between heights',
        'Maximum profit by buying and selling a share atmost twice',
        'Maximum profit by buying and selling a share at most k times',
        'Optimal Strategy for a Game', 'Palindrome Partitioning Problem',
        'Word Wrap Problem', 'Largest rectangular sub-matrix whose sum is 0',
        'Largest area rectangular sub-matrix with equal number of 1s and 0s',
        'Maximum sum rectangle in a 2D matrix',
        'Find if a string is interleaved of two other strings',
        'Coin game winner where every player has three choices',
        'Longest Increasing Subsequence', 'Maximum Sum Increasing Subsequence',
        'Longest Common Subsequence', 'Longest Palindromic Subsequence',
        'Longest Common Substring', 'Subset Sum Problem',
        'Knapsack Problem', 'Unbounded Knapsack',
        'Min Cost Path Problem', 'Minimum number of jumps to reach end',
        'Minimum cost to fill given weight in a bag',
        'Partition problem', 'Count All Palindromic Subsequence',
        'Boolean Parenthesization Problem', 'Longest Alternating Subsequence',
        'Weighted Job Scheduling', 'Count Derangements',
        'Optimal Binary Search Tree', 'Mobile Numeric Keypad Problem',
        'Maximum profit by buying and selling a share at most k times',
        'Snake and Ladders Problem', 'Clone a graph',
        'Find bridge in a graph', 'Vertex Cover Problem',
        'Chinese Postman or Route Inspection', 'Number of Triangles',
        'Minimise the cashflow', 'Two Clique Problem',
        'Rat in a maze Problem', 'Printing all solutions in N-Queen Problem',
        'Sudoku Solver', 'm Coloring Problem', 'The Knight\'s tour problem',
        'Tug of War', 'Combinational Sum', 'Partition of a set into K subsets',
        'Find the K-th Permutation Sequence',
    }
    medium_problems = {
        'Kadane\'s Algo', 'Merge Intervals', 'Next Permutation',
        'Find the Kth max and min element', 'Minimise the maximum difference between heights',
        'Minimum number of Jumps to reach end', 'Find duplicate in an array of N+1 Integers',
        'Merge 2 sorted arrays without using Extra space', 'Maximum product subarray',
        'Find longest consecutive subsequence', 'Chocolate Distribution problem',
        'Maximum profit by buying and selling a share atmost twice',
        'Rearrange the array in alternating positive and negative items',
        'Find if there is any subarray with sum equal to 0',
        'Find elements that appear more than n/k times',
        'Best time to Buy and Sell stock',
        'Spirally traverse a matrix', 'Search an element in a matrix',
        'Row with max 1s', 'Print elements in sorted order from matrix',
        'Find a specific pair in matrix', 'Rotate matrix by 90 degrees',
        'Kth smallest element in a matrix', 'Common elements in all rows',
        'Longest Repeating Subsequence', 'Longest Common Substring',
        'Word Break Problem', 'Count and Say problem',
        'Rabin-Karp Algorithm', 'Edit Distance',
        'Find next greater number with same set of digits',
        'Implement Trie', 'Find shortest unique prefix',
        'Implement a Phone Directory',
        'Implement BFS algorithm', 'Implement DFS Algo',
        'Detect Cycle in Directed Graph', 'Detect Cycle in UnDirected Graph',
        'Search in a Maze', 'Minimum Step by Knight',
        'flood fill algo', 'Clone a graph',
        'word Ladder', 'Implement Topological Sort',
        'Find whether it is possible to finish all tasks',
        'Find the no. of Islands', 'Given a sorted Dictionary of Alien Language',
        'Implement Prim\'s Algorithm', 'Graph Colouring Problem',
        'Snake and Ladders Problem', 'Check whether a graph is Bipartite',
        'Detect Negative cycle in a graph', 'Longest path in a Directed Acyclic Graph',
        'Implement a Maxheap/MinHeap', 'Sort an Array using heap',
        'Maximum of all subarrays of size k', 'k largest element in an array',
        'Kth smallest and largest element', 'Merge K sorted arrays',
        'Merge 2 Binary Max Heaps', 'Kth largest sum continuous subarrays',
        'Reorganize strings', 'Merge K Sorted Linked Lists',
        'Check if a Binary Tree is Heap', 'Connect n ropes with minimum cost',
        'Convert BST to Min Heap', 'Convert min heap to max heap',
        'Minimum sum of two numbers formed from digits',
        'Coin Change Problem', 'Knapsack Problem',
        'Binomial Coefficient Problem', 'Permutation Coefficient Problem',
        'Program for nth Catalan Number', 'Edit Distance',
        'Subset Sum Problem', 'Friends Pairing Problem',
        'Gold Mine Problem', 'Painting the Fence problem',
        'Maximum Length of Pair Chain',
        'Activity Selection Problem', 'Job Sequencing Problem',
        'Fractional Knapsack Problem', 'Greedy Algorithm to find Minimum number of Coins',
        'Minimum Platforms Problem', 'Maximum trains for which stoppage can be provided',
        'Buy Maximum Stocks if i stocks can be bought on i-th day',
        'Find the minimum and maximum amount to buy all N candies',
        'Minimum Cost to cut a board into squares',
        'Check if it is possible to survive on Island',
        'Find maximum meetings in one room', 'Maximum product subset of an array',
        'Maximize array sum after K negations', 'Maximize the sum of arr[i]*i',
        'Maximum sum of absolute difference of an array',
        'Maximize sum of consecutive differences in a circular array',
        'Minimum sum of absolute difference of pairs of two arrays',
        'Program for Shortest Job First', 'Program for Least Recently Used',
        'Smallest subset with sum greater than all other elements',
        'K Centers Problem', 'Minimum Cost of ropes',
        'Find smallest number with given number of digits and sum of digits',
        'Rearrange characters in a string such that no two adjacent are same',
        'Find maximum sum possible equal sum of three stacks',
        'Check the expression has valid or Balanced parenthesis or not',
        'Reverse a String using Stack',
        'Design a Stack that supports getMin()',
        'Find the next Greater element', 'The celebrity Problem',
        'Arithmetic Expression evaluation', 'Evaluation of Postfix expression',
        'Insert element at bottom of stack', 'Reverse a stack using recursion',
        'Sort a Stack using recursion', 'Merge Overlapping Intervals',
        'Implement Stack using Queue', 'Implement Stack using Deque',
        'Stack Permutations', 'Implement Queue using Stack',
        'Implement a Circular queue', 'LRU Cache Implementation',
        'Reverse a Queue using recursion', 'Reverse the first K elements of a queue',
        'Interleave the first half of the queue with second half',
        'Find the first circular tour that visits all Petrol Pumps',
        'Minimum time required to rot all oranges',
        'Distance of nearest cell having 1 in a binary matrix',
        'First negative integer in every window of size k',
        'Sum of minimum and maximum elements of all subarrays of size k',
        'Minimum sum of squares of character counts',
        'Queue based approach or first non-repeating character in a stream',
        'Next Smaller Element', 'Construct a trie from scratch',
        'Word Break Problem (Trie solution)',
        'Given a sequence of words, print all anagrams together',
        'Implement a Phone Directory', 'Print unique rows in a given boolean matrix',
        'Binary Trees', 'Binary Search Trees',
        'Flatten BST to sorted list', 'Find min and max value in a BST',
        'Check if a tree is a BST', 'Populate Inorder successor',
        'Find LCA of 2 nodes in a BST', 'Find Kth largest element in a BST',
        'Find Kth smallest element in a BST', 'Count pairs from 2 BST',
        'Count BST nodes that lie in a given range',
        'Deletion of a node in a BST',
        'Sum of Nodes on the Longest path from root to leaf node',
        'Find Largest subtree sum in a tree',
        'Maximum Sum of nodes such that no two are adjacent',
        'Print all K Sum paths in a Binary tree',
        'Find LCA in a Binary tree',
        'Find distance between 2 nodes',
        'Kth Ancestor of node in a Binary tree',
        'Find all Duplicate subtrees', 'Tree Isomorphism Problem',
        'Check if all leaf nodes are at same level',
        'Find if given graph is tree or not',
    }
    
    if name in hard_problems: return 'Hard'
    if name in medium_problems: return 'Medium'
    
    topic_diff = {
        'Array': 'Medium', 'Matrix': 'Medium', 'String': 'Medium',
        'Searching & Sorting': 'Medium', 'LinkedList': 'Medium',
        'Binary Trees': 'Medium', 'Binary Search Trees': 'Medium',
        'Graph': 'Medium', 'Heap': 'Medium', 'Dynamic Programming': 'Hard',
        'Greedy': 'Medium', 'BackTracking': 'Hard', 'Stacks & Queues': 'Medium',
        'Trie': 'Medium', 'Bit Manipulation': 'Easy',
        'Selection Sort': 'Easy', 'Insertion Sort': 'Easy', 'Bubble Sort': 'Easy',
    }
    
    # Easy patterns
    easy_names = {'Reverse the array', 'Find the maximum and minimum element in an array',
        'Sort an array of 0s 1s and 2s', 'Move all negative elements',
        'Cyclically rotate an array by one', 'Find union and intersection',
        'Best time to Buy and Sell stock', 'Find common elements in 3 sorted arrays',
        'Find whether an array is a subset of another array',
        'Check whether a String is Palindrome', 'Find Duplicate Characters',
        'Longest Common Prefix', 'Check if two strings are anagrams',
        'Reverse a String', 'Check if strings are rotations',
        'Check if a string is subsequence',
        'Count set bits in an integer', 'Find the two non-repeating elements',
        'Count number of bits to be flipped', 'Count total set bits',
        'Find whether a no is power of two', 'Find position of the only set bit',
        'Copy set bits in a range', 'Divide two integers',
        'Calculate square of a number', 'Power Set',
        'Implement Stack from Scratch', 'Implement Queue from Stack',
        'Implement 2 stack in an array', 'Find the middle element of a stack',
        'Implement N stacks in an Array',
        'Insertion Sort', 'Selection Sort', 'Bubble Sort',
        'Implement BFS', 'Implement DFS',
        'Binary Search', 'Search in a sorted array',
        'First and last occurrences of x', 'Find a Fixed Point',
        'Search in a Rotated Sorted Array', 'Peak element',
        'Median of two sorted arrays',
        'Find a pair with given difference', 'Find common elements',
        'Trie', 'Find a pair with given sum',
        'Add 1 to a number represented as linked list',
        'Reverse a linked list', 'Reverse a linked list in groups of size K',
        'Detect loop in a linked list',
        'Remove duplicates from an unsorted linked list',
        'Remove loop in a linked list',
        'Add two numbers represented by linked lists',
        'Find the intersection point of two linked lists',
        'Find middle element of linked list',
        'Delete middle element of linked list',
        'Nth node from end of linked list',
        'Check if linked list is palindrome',
        'Flatten a linked list',
        'Rotate a linked list',
        'Reverse a doubly linked list',
        'Merge two sorted linked lists',
        'Split a Circular Linked List into two halves',
        'Check if linked list is circular',
        'Count nodes of linked list',
        'Find length of linked list',
        'Delete a node without head reference',
        'Implement a priority queue',
        'Minimum spanning tree',
        'Find inorder successor and predecessor',
    }
    
    if name in easy_names: return 'Easy'
    return topic_diff.get(topic, 'Medium')

# Now generate solutions for ALL problems
# We need comprehensive solution templates for each topic

problems_list = []
for row_num, topic, prob_name, status, notes in raw:
    s = status_map(status)
    d = get_difficulty(topic, prob_name, row_num)
    
    problems_list.append({
        "id": slug(prob_name),
        "topic": topic,
        "problem": prob_name,
        "difficulty": d,
        "status": s,
        "notes": notes,
        "row": row_num
    })

print(f"Total problems: {len(problems_list)}")
for t in ['Array','Matrix','String','Selection Sort','Insertion Sort','Bubble Sort','Searching & Sorting','LinkedList','Binary Trees','Binary Search Trees','Graph','Heap','Dynamic Programming','Greedy','BackTracking','Stacks & Queues','Trie','Bit Manipulation']:
    count = sum(1 for p in problems_list if p['topic']==t)
    print(f"  {t}: {count}")
    
# Save raw list for next step
with open('/tmp/dsa_raw.json','w') as f:
    json.dump(problems_list, f, indent=2)
print("Saved raw list to /tmp/dsa_raw.json")
